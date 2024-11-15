import openpyxl
from datetime import datetime
import sqlite3
def otchet():
    # Подключение к базе данных SQLite
    conn = sqlite3.connect('manicur.sql')
    cursor = conn.cursor()

    today_date = datetime.now().strftime("%d.%m.%Y")  # Исправлен формат даты
    current_month = datetime.now().strftime("%m.%Y")
    print(current_month)

    # Создаем новую книгу (Excel-файл)
    workbook = openpyxl.Workbook()

    # Выбираем активный лист (по умолчанию создается один лист)
    sheet = workbook.active

    # Запрос к базе данных за день
    query_day = '''
        SELECT Код_мастера, Код_услуги, SUM(Цена) as Сумма
        FROM Прием
        WHERE Дата = ?
        GROUP BY Код_услуги
    '''

    cursor.execute(query_day, (today_date,))
    data_from_db_day = cursor.fetchall()

    # Записываем заголовок для отчета за день
    sheet['A1'] = 'Отчетность за день'
    sheet['A2'] = 'Мастер'
    sheet['B2'] = 'Услуга'
    sheet['C2'] = 'Цена'

    # Записываем данные из базы данных за день в ячейки начиная с A3
    for row_idx, row_data in enumerate(data_from_db_day, start=3):
        for col_idx, cell_value in enumerate(row_data, start=1):
            sheet.cell(row=row_idx, column=col_idx, value=cell_value)

    # Отчетность сумм за месяц
    query_month = '''
        SELECT Код_мастера, Код_услуги, SUM(Цена) as Сумма
        FROM Прием
        WHERE SUBSTR(Дата, 4, 7) = ?
        GROUP BY Код_услуги
    '''

    cursor.execute(query_month, (current_month,))
    data_from_db_month = cursor.fetchall()

    # Записываем заголовок для отчета за месяц
    sheet['D1'] = 'Отчетность за месяц'
    sheet['D2'] = 'Мастер'
    sheet['E2'] = 'Услуга'
    sheet['F2'] = 'Цена'

    # Записываем данные из базы данных за месяц в ячейки начиная с D3
    for row_idx, row_data in enumerate(data_from_db_month, start=3):
        for col_idx, cell_value in enumerate(row_data, start=4):  # Исправлен стартовый столбец
            sheet.cell(row=row_idx, column=col_idx, value=cell_value)

    # Отчетность количества услуг за день
    sheet['G1'] = 'Отчетность количества услуг за день'
    sheet['G2'] = 'Услуга'
    sheet['H2'] = 'Количество'
    sheet['I2'] = 'Цена'

    query_service_count = '''
        SELECT Код_услуги, COUNT(*) as Количество, SUM(Цена) as Сумма
        FROM Прием
        WHERE  Дата = ?
        GROUP BY Код_услуги
    '''

    cursor.execute(query_service_count, (today_date,))
    data_service_count = cursor.fetchall()

    # Записываем данные о количестве услуг и их цене в ячейки начиная с G3
    for row_idx, row_data in enumerate(data_service_count, start=3):
        sheet.cell(row=row_idx, column=7, value=row_data[0])  # Код услуги в колонку G
        sheet.cell(row=row_idx, column=8, value=row_data[1])  # Количество в колонку H
        sheet.cell(row=row_idx, column=9, value=row_data[2])  # Сумма в колонку I


    # Отчетность количества услуг за месяц
    sheet['J1'] = 'Отчетность количества услуг за месяц'
    sheet['J2'] = 'Услуга'
    sheet['K2'] = 'Количество'
    sheet['L2'] = 'Цена'

    query_service_count = '''
        SELECT Код_услуги, COUNT(*) as Количество, SUM(Цена) as Сумма
        FROM Прием
        WHERE SUBSTR(Дата, 4, 7) = ?
        GROUP BY Код_услуги
    '''
    cursor.execute(query_service_count, (current_month,))
    data_service_count = cursor.fetchall()
    # Записываем данные о количестве услуг и их цене в ячейки начиная с G3
    for row_idx, row_data in enumerate(data_service_count, start=3):
        sheet.cell(row=row_idx, column=10, value=row_data[0])  # Код услуги в колонку G
        sheet.cell(row=row_idx, column=11, value=row_data[1])  # Количество в колонку H
        sheet.cell(row=row_idx, column=12, value=row_data[2])  # Сумма в колонку I


    # Сумма доходов за день
    sheet['M1'] = 'Сумма доходов за день'
    sheet['M2'] = 'Услуга'
    sheet['N2'] = 'Сумма'

    query_day_income = '''
        SELECT Код_услуги, SUM(Цена) as Сумма
        FROM Прием
        WHERE Дата = ?
        GROUP BY Код_услуги
    '''

    cursor.execute(query_day_income, (today_date,))
    data_day_income = cursor.fetchall()

    # Записываем данные из базы данных о сумме доходов за день в ячейки начиная с J3
    for row_idx, row_data in enumerate(data_day_income, start=3):
        sheet.cell(row=row_idx, column=13, value=row_data[0])  # Код услуги в колонку J
        sheet.cell(row=row_idx, column=14, value=row_data[1])  # Сумма в колонку K


    # Сохраняем файл
    workbook.save('отчет.xlsx')

    # Закрываем подключение к базе данных
    conn.close()